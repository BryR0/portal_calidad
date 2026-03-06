from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
import io
import json
from datetime import datetime, date

from flask import render_template, redirect, url_for, flash, request, Response, jsonify
from flask_login import login_required, current_user
from sqlalchemy import inspect, or_

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.routes import nc_bp
from app.models.nc import NonConformity, Provider
from app.models.catalogs import Empresa, Marca, Seccion, ClasificacionNC, CatalogItem
from app.models.maestras import ReporteMaestro, RSBodega, RSInvPorDia, Suplidor
# Import local en las funciones que requieran 'db'
from app.data.catalogs import (
    DETECTADA, DESVIACION, GRAVEDAD, MOTIVO,
    PROVEEDOR_TIPO, PROVEEDOR_ABBREV,
    GRAVEDAD_COLOR, STATUS_COLOR,
)


def _has_table(table_name):
    from app import db
    return inspect(db.engine).has_table(table_name)


def _catalog_empresas():
    if _has_table('rsi_bodegas'):
        return [f"{b.bodega} - {b.nombre}" for b in RSBodega.query
                .order_by(RSBodega.nombre).all()]
    if _has_table('empresas'):
        return [f"{e.codigo} - {e.name}" if e.codigo else e.name for e in Empresa.query
                .filter_by(is_active=True)
                .order_by(Empresa.name).all()]
    return []


def _catalog_secciones():
    secciones_por_tipo = {}
    secciones_por_rimith = {}

    rows = (Seccion.query
            .filter_by(is_active=True)
            .order_by(Seccion.sort_order, Seccion.name)
            .all())

    for s in rows:
        if s.provider_tipo == 'Rimith':
            dept = s.rimith_dept or 'Produccion'
            secciones_por_rimith.setdefault(dept, []).append(s.name)
            continue
        if s.provider_tipo:
            secciones_por_tipo.setdefault(s.provider_tipo, []).append(s.name)

    return {
        'flat': [s.name for s in rows],
        'by_tipo': secciones_por_tipo,
        'rimith': secciones_por_rimith,
    }


def _catalog_marcas():
    from app import db

    if _has_table('reporte_maestro'):
        marcas_db = db.session.query(ReporteMaestro.marcanomb).distinct().filter(
            ReporteMaestro.marcanomb.isnot(None),
            ReporteMaestro.marcanomb != '',
            ReporteMaestro.rechazado == 0
        ).order_by(ReporteMaestro.marcanomb).all()
        return [m[0] for m in marcas_db]

    if _has_table('marcas'):
        return [m.name for m in Marca.query
                .filter_by(is_active=True)
                .order_by(Marca.name).all()]

    return []


def _buscar_proveedores(q, tipo=''):
    if _has_table('providers'):
        query = Provider.query.filter_by(is_active=True)
        if tipo:
            query = query.filter(Provider.provider_type == tipo)
        results = (query
                   .filter(Provider.name.ilike(f'%{q}%'))
                   .order_by(Provider.name)
                   .limit(30)
                   .all())
        if results:
            return [{
                'id': r.id,
                'name': r.name,
                'label': f'{r.id} - {r.name}',
            } for r in results]

    if _has_table('rsi_cliente_suplidor'):
        results = Suplidor.query.filter(
            Suplidor.activo == True,
            or_(
                Suplidor.cliente_suplidor.ilike(f'%{q}%'),
                Suplidor.nombre.ilike(f'%{q}%'),
                Suplidor.ruc.ilike(f'%{q}%')
            )
        ).limit(30).all()
        return [{
            'id': None,
            'name': r.nombre,
            'label': f'{r.cliente_suplidor} - {r.nombre}',
        } for r in results]

    return []


def _resolve_provider_selection(provider_type, raw_name, raw_id=''):
    provider_name = (raw_name or '').strip()
    normalized_name = provider_name.split(
        ' - ', 1)[1].strip() if ' - ' in provider_name else provider_name

    if raw_id and str(raw_id).isdigit() and _has_table('providers'):
        provider = Provider.query.filter_by(
            id=int(raw_id), is_active=True).first()
        if provider and (not provider_type or provider.provider_type == provider_type):
            return provider.id, provider.name

    if normalized_name and _has_table('providers'):
        query = Provider.query.filter_by(is_active=True, name=normalized_name)
        if provider_type:
            query = query.filter(Provider.provider_type == provider_type)
        provider = query.order_by(Provider.id).first()
        if provider:
            return provider.id, provider.name

    return None, normalized_name


def _parse_float(raw_value, default=0.0):
    value = (str(raw_value).strip() if raw_value is not None else '')
    if not value:
        return default
    try:
        return float(value.replace(',', ''))
    except ValueError:
        return default


def _extract_bodega_code(empresa_value):
    empresa = (empresa_value or '').strip()
    if not empresa:
        return ''
    if ' - ' in empresa:
        return empresa.split(' - ', 1)[0].strip()
    if _has_table('empresas'):
        row = Empresa.query.filter_by(name=empresa, is_active=True).first()
        if row and row.codigo:
            return row.codigo.strip()
    return empresa


def _lookup_branch_inventory(item_code, empresa_value):
    if not item_code or not _has_table('rsi_inv_pordia'):
        return {
            'bodega': '',
            'branch_total_units': 0.0,
            'branch_total_usd': 0.0,
            'unit_price': 0.0,
        }

    bodega_code = _extract_bodega_code(empresa_value)
    if not bodega_code:
        return {
            'bodega': '',
            'branch_total_units': 0.0,
            'branch_total_usd': 0.0,
            'unit_price': 0.0,
        }

    row = (RSInvPorDia.query
           .filter_by(bodega=bodega_code, item=item_code)
           .order_by(RSInvPorDia.fecha.desc())
           .first())

    if not row:
        return {
            'bodega': bodega_code,
            'branch_total_units': 0.0,
            'branch_total_usd': 0.0,
            'unit_price': 0.0,
        }

    units = _parse_float(row.cantidad)
    unit_price = _parse_float(row.precio)
    return {
        'bodega': bodega_code,
        'branch_total_units': units,
        'branch_total_usd': round(units * unit_price, 2),
        'unit_price': unit_price,
    }


def _resolve_master_product(raw_product_code='', raw_barcode=''):
    if not _has_table('reporte_maestro'):
        return None

    product_code = (raw_product_code or '').strip()
    barcode = (raw_barcode or '').strip()

    if product_code:
        product = ReporteMaestro.query.filter(
            ReporteMaestro.itemcod == product_code,
            ReporteMaestro.rechazado == 0
        ).first()
        if product:
            return product

    if barcode:
        product = ReporteMaestro.query.filter(
            ReporteMaestro.itemcodbarra == barcode,
            ReporteMaestro.rechazado == 0
        ).first()
        if product:
            return product

    if barcode and barcode != product_code:
        product = ReporteMaestro.query.filter(
            ReporteMaestro.itemcod == barcode,
            ReporteMaestro.rechazado == 0
        ).first()
        if product:
            return product

    if product_code and product_code != barcode:
        product = ReporteMaestro.query.filter(
            ReporteMaestro.itemcodbarra == product_code,
            ReporteMaestro.rechazado == 0
        ).first()
        if product:
            return product

    return None


# ── Catalogs helper (queries DB for user-managed catalogs) ────────────────────
def _catalogs():
    # Small static lists (rarely change)
    ctx = dict(
        cat_detectada=DETECTADA,
        cat_desviacion=DESVIACION,
        cat_gravedad=GRAVEDAD,
        cat_motivo=MOTIVO,
        cat_proveedor_tipo=PROVEEDOR_TIPO,
        gravedad_color=GRAVEDAD_COLOR,
        status_color=STATUS_COLOR,
    )

    # DB-driven catalogs
    ctx['cat_causa'] = [c.value for c in CatalogItem.query
                        .filter_by(category='CAUSA', is_active=True)
                        .order_by(CatalogItem.sort_order).all()]
    ctx['cat_prerrequisito'] = [c.value for c in CatalogItem.query
                                .filter_by(category='PRERREQUISITO', is_active=True)
                                .order_by(CatalogItem.sort_order).all()]
    ctx['cat_evaluacion'] = [c.value for c in CatalogItem.query
                             .filter_by(category='EVALUACION', is_active=True)
                             .order_by(CatalogItem.sort_order).all()]
    ctx['cat_clasificacion'] = [c.name for c in ClasificacionNC.query
                                .filter_by(is_active=True)
                                .order_by(ClasificacionNC.sort_order).all()]

    ctx['cat_empresas'] = _catalog_empresas()

    ctx['cat_marcas'] = _catalog_marcas()

    # Cascading dropdowns JSON ─────────────────────────────────────────────
    # Ya no cargamos los 120,000 proveedores en HTML. Se usa AJAX.
    ctx['cat_proveedores_json'] = {}

    secciones = _catalog_secciones()
    ctx['cat_secciones_json'] = secciones['by_tipo']
    ctx['cat_rimith_secciones_json'] = secciones['rimith']
    ctx['cat_rimith_depts'] = list(secciones['rimith'].keys())
    ctx['cat_secciones'] = secciones['flat']

    return ctx


# ── NC number generator ───────────────────────────────────────────────────────
def _generate_nc_number(provider_type, detected_date):
    prefix = detected_date.strftime('%d%m%Y')
    abbrev = PROVEEDOR_ABBREV.get(provider_type, 'XX')
    count = NonConformity.query.filter(
        NonConformity.nc_number.like(f'{prefix}-%')).count() + 1
    return f'{prefix}-{count:03d}-{abbrev}'


# ── Query builder (filters) ───────────────────────────────────────────────────
def _build_query(args):
    q = NonConformity.query
    status = args.get('status')
    if status:
        q = q.filter(NonConformity.status.in_(['Cerrada', 'Cerrado'])
                     if status == 'Cerrada' else NonConformity.status == status)
    if args.get('gravity'):
        q = q.filter_by(gravity=args['gravity'])
    section = args.get('section', '').strip()
    if section:
        q = q.filter(NonConformity.section.ilike(f'%{section}%'))
    search = args.get('q', '').strip()
    if search:
        q = q.filter(or_(
            NonConformity.nc_number.ilike(f'%{search}%'),
            NonConformity.title.ilike(f'%{search}%'),
            NonConformity.product_code.ilike(f'%{search}%'),
            NonConformity.barcode.ilike(f'%{search}%'),
            NonConformity.product_affected.ilike(f'%{search}%'),
            NonConformity.description.ilike(f'%{search}%'),
        ))
    if args.get('date_from'):
        try:
            q = q.filter(NonConformity.date_detected >=
                         datetime.strptime(args['date_from'], '%Y-%m-%d').date())
        except ValueError:
            pass
    if args.get('date_to'):
        try:
            q = q.filter(NonConformity.date_detected <=
                         datetime.strptime(args['date_to'], '%Y-%m-%d').date())
        except ValueError:
            pass
    return q


def _can_edit_nc(nc):
    return nc.status == 'Abierta' or current_user.is_calidad()


def _apply_form_to_nc(nc, form, *, is_new=False):
    date_str = form.get('date_detected')
    detected_date = (datetime.strptime(date_str, '%Y-%m-%d').date()
                     if date_str else date.today())
    time_str = form.get('time_detected') or '00:00'
    detected_time = datetime.strptime(time_str, '%H:%M').time()

    provider_type = form.get('provider_tipo', '').strip()
    raw_nc_number = form.get('nc_number', '').strip()
    if is_new:
        nc.nc_number = raw_nc_number or _generate_nc_number(
            provider_type, detected_date)
    elif raw_nc_number:
        nc.nc_number = raw_nc_number

    exp_date = None
    exp_str = form.get('expiration_date', '').strip()
    if exp_str:
        try:
            exp_date = datetime.strptime(exp_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    provider_name = form.get('provider_name', '').strip()
    prov_id, provider_name = _resolve_provider_selection(
        provider_type,
        provider_name,
        form.get('provider_id', '').strip(),
    )
    empresa = form.get('empresa', '').strip() or None
    rimith_dept = form.get('rimith_dept', '').strip(
    ) if provider_type == 'Rimith' else None
    section = form.get('section', '').strip(
    ) if provider_type == 'Rimith' else None
    raw_product_code = form.get('product_code', '').strip(
    ) or form.get('itemcod', '').strip()
    raw_barcode = form.get('barcode', '').strip()
    matched_product = _resolve_master_product(raw_product_code, raw_barcode)

    if matched_product:
        product_code = (matched_product.itemcod or '').strip()
        barcode = (matched_product.itemcodbarra or '').strip() or None
        product_affected = (matched_product.itemnombre or form.get(
            'product_affected', '').strip()) or None
        marca = (matched_product.marcanomb or form.get(
            'marca', '').strip()) or None
    else:
        product_code = raw_product_code or None
        barcode = raw_barcode or None
        product_affected = form.get('product_affected', '').strip() or None
        marca = form.get('marca', '').strip() or None

    branch_inventory = _lookup_branch_inventory(product_code, empresa)

    if provider_type == 'Rimith' and (not rimith_dept or not section):
        raise ValueError('Rimith requiere departamento y seccion.')

    nc.date_detected = detected_date
    nc.time_detected = detected_time
    nc.detected_by_type = form.get('detected_by_type')
    nc.empresa = empresa
    nc.provider_tipo = provider_type
    nc.provider_id = prov_id
    nc.provider_name = provider_name or None
    nc.rimith_dept = rimith_dept or None
    nc.section = section or None
    nc.deviation_type = form.get('deviation_type')
    nc.gravity = form.get('gravity')
    nc.cause = form.get('cause')
    nc.motive = form.get('motive')
    nc.prerequisite = form.get('prerequisite')
    nc.title = form.get('title', '').strip()
    nc.description = form.get('description', '').strip()
    nc.product_code = product_code
    nc.product_affected = product_affected
    nc.marca = marca
    nc.barcode = barcode
    nc.lot = form.get('lot', '').strip() or None
    nc.expiration_date = exp_date
    nc.total_units = _parse_float(form.get('total_units'))
    nc.total_usd = _parse_float(form.get('total_usd'))
    nc.branch_total_units = branch_inventory['branch_total_units']
    nc.branch_total_usd = branch_inventory['branch_total_usd']
    nc.affects_consumer = bool(form.get('affects_consumer'))
    nc.affects_inocuity = bool(form.get('affects_inocuity'))
    nc.affects_quality = bool(form.get('affects_quality'))
    nc.nc_product_status = form.get('nc_product_status', '').strip() or None
    nc.evaluations = form.get('evaluation_type', '').strip() or None
    nc.evaluation_description = form.get(
        'evaluation_description', '').strip() or None
    nc.classification = form.get('classification', '').strip() or None
    nc.action_1 = form.get('action_1', '').strip() or None
    nc.action_2 = form.get('action_2', '').strip() or None
    nc.action_3 = form.get('action_3', '').strip() or None
    if 'corrective_action' in form:
        nc.corrective_action = form.get(
            'corrective_action', '').strip() or None

    if is_new:
        nc.status = 'Abierta'

    return nc


# ── LIST ──────────────────────────────────────────────────────────────────────
@nc_bp.route('/')
@login_required
def list_nc():
    page = request.args.get('page', 1, type=int)
    pagination = _build_query(request.args).order_by(
        NonConformity.date_detected.desc()
    ).paginate(page=page, per_page=25, error_out=False)
    return render_template('nc/list.html', ncs=pagination.items,
                           pagination=pagination, filters=request.args,
                           **_catalogs())


# ── CREATE ────────────────────────────────────────────────────────────────────
@nc_bp.route('/nueva', methods=['GET', 'POST'])
@login_required
def create_nc():
    if request.method == 'POST':
        try:
            new_nc = _apply_form_to_nc(
                NonConformity(), request.form, is_new=True)
            from app import db
            db.session.add(new_nc)
            db.session.commit()
            flash(f'NC {new_nc.nc_number} registrada exitosamente.', 'success')
            return redirect(url_for('nc.view_nc', id=new_nc.id))
        except Exception as exc:
            from app import db
            db.session.rollback()
            flash(f'Error al registrar NC: {exc}', 'error')

    return render_template('nc/create.html', today=date.today().isoformat(),
                           is_edit=False, nc=None, **_catalogs())


@nc_bp.route('/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def edit_nc(id):
    nc = NonConformity.query.get_or_404(id)
    if not _can_edit_nc(nc):
        flash('No tienes permisos para editar esta NC.', 'error')
        return redirect(url_for('nc.view_nc', id=id))

    if request.method == 'POST':
        try:
            _apply_form_to_nc(nc, request.form, is_new=False)
            from app import db
            db.session.commit()
            flash(f'NC {nc.nc_number} actualizada exitosamente.', 'success')
            return redirect(url_for('nc.view_nc', id=nc.id))
        except Exception as exc:
            from app import db
            db.session.rollback()
            flash(f'Error al actualizar NC: {exc}', 'error')

    return render_template(
        'nc/create.html',
        today=(nc.date_detected.isoformat()
               if nc.date_detected else date.today().isoformat()),
        is_edit=True,
        nc=nc,
        **_catalogs(),
    )


# ── VIEW ──────────────────────────────────────────────────────────────────────
@nc_bp.route('/<int:id>')
@login_required
def view_nc(id):
    nc = NonConformity.query.get_or_404(id)
    return render_template('nc/view.html', nc=nc, **_catalogs())


# ── CLOSE ─────────────────────────────────────────────────────────────────────
@nc_bp.route('/<int:id>/cerrar', methods=['POST'])
@login_required
def close_nc(id):
    if not current_user.is_calidad():
        flash('No tienes permisos para cerrar NCs.', 'error')
        return redirect(url_for('nc.view_nc', id=id))
    nc = NonConformity.query.get_or_404(id)
    nc.status = 'Cerrada'
    nc.closure_date = date.today()
    nc.closed_by_id = current_user.id
    nc.action_1 = request.form.get('action_1', nc.action_1)
    nc.action_2 = request.form.get('action_2', nc.action_2)
    nc.action_3 = request.form.get('action_3', nc.action_3)
    nc.corrective_action = request.form.get('corrective_action')
    nc.evaluation_description = request.form.get(
        'evaluation_description', nc.evaluation_description)
    from app import db
    db.session.commit()
    flash(f'NC {nc.nc_number} cerrada exitosamente.', 'success')
    return redirect(url_for('nc.view_nc', id=id))


# ── PRINT ─────────────────────────────────────────────────────────────────────
@nc_bp.route('/<int:id>/imprimir')
@login_required
def print_nc(id):
    nc = NonConformity.query.get_or_404(id)
    return render_template('nc/print.html', nc=nc, now=datetime.now())


# ── PRINT LIST ────────────────────────────────────────────────────────────────
@nc_bp.route('/imprimir-lista')
@login_required
def print_list_nc():
    ncs = _build_query(request.args).order_by(
        NonConformity.date_detected.desc()).all()
    return render_template('nc/print_list.html', ncs=ncs,
                           filters=request.args, now=datetime.now())


# ── EXPORT LIST → XLSX ────────────────────────────────────────────────────────
@nc_bp.route('/exportar-excel')
@login_required
def export_excel():
    ncs = _build_query(request.args).order_by(
        NonConformity.date_detected.desc()).all()
    buf = io.BytesIO()
    _build_list_workbook(ncs).save(buf)
    buf.seek(0)
    fname = f'NC_Reporte_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})


# ── EXPORT SINGLE → XLSX ──────────────────────────────────────────────────────
@nc_bp.route('/<int:id>/exportar-excel')
@login_required
def export_nc_excel(id):
    nc = NonConformity.query.get_or_404(id)
    buf = io.BytesIO()
    _build_single_workbook(nc).save(buf)
    buf.seek(0)
    fname = f'NC_{nc.nc_number.replace("/", "-")}.xlsx'
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})


# ── API ENDPOINTS (AJAX) ──────────────────────────────────────────────────────
@nc_bp.route('/api/producto/buscar')
@login_required
def api_buscar_producto():
    """
    Busca un producto por código o código de barras en ReporteMaestro.
    Retorna JSON para autocompletar la sección de Producto NC.
    """
    q = request.args.get('q', '').strip()
    empresa = request.args.get('empresa', '').strip()
    if not q:
        return jsonify({'success': False, 'message': 'Debe ingresar un término de búsqueda'}), 400

    from sqlalchemy import or_

    def _serialize_product(product):
        inventory = _lookup_branch_inventory(product.itemcod, empresa)
        return {
            'itemcod': product.itemcod,
            'barcode': product.itemcodbarra or '',
            'nombre': product.itemnombre or '',
            'marca': product.marcanomb or '',
            'branch_total_units': inventory['branch_total_units'],
            'branch_total_usd': inventory['branch_total_usd'],
            'branch_bodega': inventory['bodega'],
            'unit_price': inventory['unit_price'],
        }

    # Intentar búsqueda exacta primero
    exact_match = ReporteMaestro.query.filter(
        or_(
            ReporteMaestro.itemcod == q,
            ReporteMaestro.itemcodbarra == q
        ),
        ReporteMaestro.rechazado == 0
    ).first()

    if exact_match:
        return jsonify({
            'success': True,
            'data': [_serialize_product(exact_match)]
        })

    # Si no hay match exacto, hacer LIKE
    results = ReporteMaestro.query.filter(
        or_(
            ReporteMaestro.itemcod.ilike(f'%{q}%'),
            ReporteMaestro.itemcodbarra.ilike(f'%{q}%')
        ),
        ReporteMaestro.rechazado == 0
    ).limit(10).all()

    if results:
        return jsonify({
            'success': True,
            'data': [_serialize_product(r) for r in results]
        })

    return jsonify({'success': False, 'message': 'No se encontraron productos'}), 404


@nc_bp.route('/api/proveedor/buscar')
@login_required
def api_buscar_proveedor():
    """
    Busca proveedores sin mezclar empresas o bodegas en el campo proveedor.
    """
    q = request.args.get('q', '').strip()
    tipo = request.args.get('tipo', '').strip()
    if not q or len(q) < 2:
        return jsonify([])
    return jsonify(_buscar_proveedores(q, tipo))

# ─────────────────────────────────────────────────────────────────────────────
# Excel builders
# ─────────────────────────────────────────────────────────────────────────────


def _border(color='D1D5DB'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS  compartidos
# ══════════════════════════════════════════════════════════════════════════════

def _fill(c):
    return PatternFill("solid", fgColor=c)


def _font(size=10, bold=False, color="0F172A", name="Calibri", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_all(color="E2E8F0", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _border_bottom(color="E2E8F0", style="thin"):
    return Border(bottom=Side(style=style, color=color))


def _border_top(color, style="medium"):
    return Border(top=Side(style=style, color=color))


# Paleta unificada
C_DARK_BLUE = "1E3A8A"
C_MID_BLUE = "2563EB"
C_ACCENT = "F6AD55"
C_WHITE = "FFFFFF"
C_SOFT = "EEF2FF"
C_GRAY_BG = "F8FAFC"
C_GRAY_LINE = "E2E8F0"
C_TEXT_DARK = "0F172A"

C_CONSUMER_BG = "FFFBEB"
C_CONSUMER_BD = "FCD34D"
C_CONSUMER_TXT = "92400E"
C_INOCUITY_BG = "FFF1F2"
C_INOCUITY_BD = "FCA5A5"
C_INOCUITY_TXT = "991B1B"
C_QUALITY_BG = "FFF7ED"
C_QUALITY_BD = "FDBA74"
C_QUALITY_TXT = "9A3412"
C_NO_BG = "E2E8F0"
C_NO_TXT = "475569"

ALIGN_C = _align("center")
ALIGN_L = _align("left")
ALIGN_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ══════════════════════════════════════════════════════════════════════════════
#  _build_single_workbook
# ══════════════════════════════════════════════════════════════════════════════

def _build_single_workbook(nc):
    wb = Workbook()
    ws = wb.active
    ws.title = (nc.nc_number or 'NC')[:28]
    ws.sheet_view.showGridLines = False

    col_widths = {"A": 3, "B": 20, "C": 22, "D": 20, "E": 22,
                  "F": 20, "G": 22, "H": 16, "I": 22}
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    COL_S = 2
    COL_E = 9

    # ── helpers locales ────────────────────────────────────────────────────
    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)

    def cell(r, c, value="", fill=None, font=None, align=None, border=None):
        cel = ws.cell(row=r, column=c, value=value)
        if fill:
            cel.fill = fill
        if font:
            cel.font = font
        if align:
            cel.alignment = align
        if border:
            cel.border = border
        return cel

    def section_header(row, title):
        ws.row_dimensions[row].height = 20
        merge(row, COL_S, row, COL_E)
        cell(row, COL_S, title,
             fill=_fill(C_DARK_BLUE),
             font=_font(10, bold=True, color=C_WHITE),
             align=_align("left"),
             border=_border_bottom(C_ACCENT, "medium"))

    def label_val(row, col_l, col_v, label, value, span=1):
        cell(row, col_l, label,
             fill=_fill(C_SOFT),
             font=_font(9, bold=True, color=C_DARK_BLUE),
             align=ALIGN_L,
             border=_border_all(C_GRAY_LINE))
        if span > 1:
            merge(row, col_v, row, col_v + span - 1)
        cell(row, col_v, str(value) if value else "",
             fill=_fill(C_WHITE),
             font=_font(10, color=C_TEXT_DARK),
             align=ALIGN_L,
             border=_border_all(C_GRAY_LINE))

    def full_row_label_val(row, label, value, height=15):
        ws.row_dimensions[row].height = height
        cell(row, COL_S, label,
             fill=_fill(C_SOFT),
             font=_font(9, bold=True, color=C_DARK_BLUE),
             align=ALIGN_L,
             border=_border_all(C_GRAY_LINE))
        merge(row, COL_S + 1, row, COL_E)
        cell(row, COL_S + 1, str(value) if value else "",
             fill=_fill(C_WHITE),
             font=_font(10, color=C_TEXT_DARK),
             align=ALIGN_L,
             border=_border_all(C_GRAY_LINE))

    def text_block(r1, r2, value, label=None):
        if label:
            merge(r1, COL_S, r2, COL_S)
            cell(r1, COL_S, label,
                 fill=_fill(C_SOFT),
                 font=_font(9, bold=True, color=C_DARK_BLUE),
                 align=ALIGN_C,
                 border=_border_all(C_GRAY_LINE))
            merge(r1, COL_S + 1, r2, COL_E)
            cell(r1, COL_S + 1, str(value) if value else "",
                 fill=_fill(C_WHITE),
                 font=_font(10, color=C_TEXT_DARK),
                 align=ALIGN_LT,
                 border=_border_all(C_GRAY_LINE))
        else:
            merge(r1, COL_S, r2, COL_E)
            cell(r1, COL_S, str(value) if value else "",
                 fill=_fill(C_WHITE),
                 font=_font(10, color=C_TEXT_DARK),
                 align=ALIGN_LT,
                 border=_border_all(C_GRAY_LINE))
        for r in range(r1, r2 + 1):
            ws.row_dimensions[r].height = 22

    # ── ROW 1  Título ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    merge(1, COL_S, 1, COL_E)
    cell(1, COL_S,
         "REPORTE DE DESVIACIONES  /  NON CONFORMING REPORT",
         fill=_fill(C_DARK_BLUE),
         font=_font(16, bold=True, color=C_WHITE),
         align=ALIGN_C,
         border=_border_bottom(C_ACCENT, "medium"))

    # ── ROW 2  Sub-header ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    merge(2, COL_S, 2, 5)
    cell(2, COL_S,
         "GRUPO RIBA SMITH  —  GESTIÓN DE CALIDAD",
         fill=_fill(C_DARK_BLUE),
         font=_font(10, bold=True, color=C_WHITE),
         align=_align("left"))
    merge(2, 6, 2, COL_E)
    cell(2, 6,
         f"Código: RS-SM-GC-REG-0001    |    Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
         fill=_fill(C_DARK_BLUE),
         font=_font(9, color="BFDBFE"),
         align=_align("right"))

    # ── ROW 3  Separador naranja ───────────────────────────────────────────
    ws.row_dimensions[3].height = 4
    for c in range(COL_S, COL_E + 1):
        ws.cell(row=3, column=c).border = _border_bottom(C_ACCENT, "medium")

    # ── ROWS 4-5  Meta cards ──────────────────────────────────────────────
    meta_cards = [
        ("N° NC",    nc.nc_number),
        ("Fecha",    nc.date_detected.strftime(
            '%d/%m/%Y') if nc.date_detected else ""),
        ("Hora",     nc.time_detected.strftime(
            '%H:%M') if nc.time_detected else ""),
        ("Gravedad", nc.gravity),
        ("Estado",   nc.status),
    ]
    card_ranges = [(2, 2), (3, 4), (5, 5), (6, 7), (8, 9)]

    ws.row_dimensions[4].height = 13
    ws.row_dimensions[5].height = 24

    for (rs, re), (lbl, val) in zip(card_ranges, meta_cards):
        if rs < re:
            merge(4, rs, 4, re)
            merge(5, rs, 5, re)
        c4 = ws.cell(row=4, column=rs, value=lbl)
        c4.fill = _fill(C_SOFT)
        c4.font = _font(8, bold=True, color=C_MID_BLUE)
        c4.alignment = ALIGN_C
        c4.border = _border_all("C7D2FE")
        c5 = ws.cell(row=5, column=rs, value=val)
        c5.fill = _fill(C_WHITE)
        c5.font = _font(13, bold=True, color=C_DARK_BLUE)
        c5.alignment = ALIGN_C
        c5.border = _border_all("C7D2FE")

    # ── ROW 6  Separador ──────────────────────────────────────────────────
    ws.row_dimensions[6].height = 6
    for c in range(COL_S, COL_E + 1):
        ws.cell(row=6, column=c).border = _border_bottom(C_GRAY_LINE)

    r = 7

    # ── S1  IDENTIFICACIÓN Y ORIGEN ───────────────────────────────────────
    section_header(r, "  1.  IDENTIFICACIÓN Y ORIGEN")
    r += 1

    ws.row_dimensions[r].height = 30
    label_val(r, 2, 3, "EMPRESA",       nc.empresa)
    label_val(r, 4, 5, "N° NC",         nc.nc_number)
    label_val(r, 6, 7, "DETECTADA",     nc.detected_by_type)
    label_val(r, 8, 9, "TIPO DESV.",    nc.deviation_type)
    r += 1

    ws.row_dimensions[r].height = 30
    label_val(r, 2, 3, "PROVEEDOR",     nc.provider_name)
    label_val(r, 4, 5, "T. PROVEEDOR",  nc.provider_tipo)
    label_val(r, 6, 7, "DEPTO RIMITH",  nc.rimith_dept)
    label_val(r, 8, 9, "ESTADO",        nc.status)
    r += 1

    full_row_label_val(r, "SECCIÓN", nc.section)
    r += 1

    # ── S2  HALLAZGO ──────────────────────────────────────────────────────
    section_header(r, "  2.  HALLAZGO")
    r += 1
    full_row_label_val(r, "TÍTULO", nc.title)
    r += 1
    text_block(r, r + 2, nc.description, label="DESCRIPCIÓN")
    r += 3

    # ── S3  PRODUCTO NO CONFORME ──────────────────────────────────────────
    section_header(r, "  3.  PRODUCTO NO CONFORME")
    r += 1
    full_row_label_val(r, "PRODUCTO", nc.product_affected)
    r += 1

    ws.row_dimensions[r].height = 15
    label_val(r, 2, 3, "MARCA",      nc.marca)
    label_val(r, 4, 5, "LOTE",       nc.lot)
    label_val(r, 6, 7, "CÓD. ÍTEM",  nc.product_code)
    label_val(r, 8, 9, "CÓD. BARRA", nc.barcode)
    r += 1

    # FIX: altura aumentada para F. VENCE / ESTADO PROD.
    ws.row_dimensions[r].height = 30
    label_val(r, 2, 3, "F. VENCE",    nc.expiration_date.strftime(
        '%d/%m/%Y') if nc.expiration_date else "")
    label_val(r, 4, 5, "ESTADO PROD.", nc.nc_product_status)
    for cc in range(6, 10):
        ws.cell(row=r, column=cc).fill = _fill(C_GRAY_BG)
        ws.cell(row=r, column=cc).border = _border_all(C_GRAY_LINE)
    r += 1

    # Métricas
    metric_data = [
        ("INV. SUCURSAL",  f'{nc.branch_total_units or 0:.2f}'),
        ("VALOR SUCURSAL", f'${nc.branch_total_usd or 0:,.2f}'),
        ("UNIDADES NC",    f'{nc.total_units or 0:.2f}'),
        ("TOTAL NC  $",    f'${nc.total_usd or 0:,.2f}'),
    ]
    ws.row_dimensions[r].height = 13
    ws.row_dimensions[r + 1].height = 28
    for (lbl, val), cs in zip(metric_data, [2, 4, 6, 8]):
        merge(r,     cs, r,     cs + 1)
        merge(r + 1, cs, r + 1, cs + 1)
        lc = ws.cell(row=r, column=cs, value=lbl)
        lc.fill = _fill(C_SOFT)
        lc.font = _font(8, bold=True, color=C_MID_BLUE)
        lc.alignment = ALIGN_C
        lc.border = _border_all("C7D2FE")
        vc = ws.cell(row=r + 1, column=cs, value=val)
        vc.fill = _fill("EFF6FF")
        vc.font = _font(14, bold=True, color=C_DARK_BLUE)
        vc.alignment = ALIGN_C
        vc.border = _border_all("BFDBFE")
    r += 2

    # ── S4  CLASIFICACIÓN Y ANÁLISIS ──────────────────────────────────────
    section_header(r, "  4.  CLASIFICACIÓN Y ANÁLISIS")
    r += 1

    # FIX: altura aumentada para la fila de causa/motivo/evaluación
    ws.row_dimensions[r].height = 30
    label_val(r, 2, 3, "CAUSA",         nc.cause)
    label_val(r, 4, 5, "MOTIVO",        nc.motive)
    label_val(r, 6, 7, "EVALUACIÓN",    nc.evaluations)
    label_val(r, 8, 9, "CLASIFICACIÓN", nc.classification)
    r += 1

    full_row_label_val(r, "PREREQUISITO", nc.prerequisite)
    r += 1
    text_block(r, r + 2, nc.evaluation_description, label="ANÁLISIS")
    r += 3

    # ── S5  IMPACTO ───────────────────────────────────────────────────────
    section_header(r, "  5.  IMPACTO")
    r += 1

    impact_rows = [
        (nc.affects_consumer, "Afecta satisfacción del consumidor final",
         C_CONSUMER_BG, C_CONSUMER_BD, C_CONSUMER_TXT),
        (nc.affects_inocuity, "Afecta inocuidad directa",
         C_INOCUITY_BG, C_INOCUITY_BD, C_INOCUITY_TXT),
        (nc.affects_quality,  "Afecta calidad directa",
         C_QUALITY_BG,  C_QUALITY_BD,  C_QUALITY_TXT),
    ]
    for flag, desc, bg, bd, txt in impact_rows:
        ws.row_dimensions[r].height = 18
        pill_bg = bg if flag else C_NO_BG
        pill_txt = txt if flag else C_NO_TXT
        cell(r, 2, "SÍ" if flag else "NO",
             fill=_fill(pill_bg),
             font=_font(10, bold=True, color=pill_txt),
             align=ALIGN_C,
             border=Border(
                 left=Side(style="medium", color=bd if flag else C_GRAY_LINE),
                 right=Side(style="thin",  color=bd if flag else C_GRAY_LINE),
                 top=Side(style="thin",    color=bd if flag else C_GRAY_LINE),
                 bottom=Side(style="thin", color=bd if flag else C_GRAY_LINE),
             ))
        merge(r, 3, r, COL_E)
        cell(r, 3, desc,
             fill=_fill(bg if flag else C_GRAY_BG),
             font=_font(10, bold=bool(flag), color=C_TEXT_DARK),
             align=ALIGN_L,
             border=_border_all(bd if flag else C_GRAY_LINE))
        r += 1

    # ── S6  ACCIONES INMEDIATAS ───────────────────────────────────────────
    section_header(r, "  6.  ACCIONES INMEDIATAS  /  IMMEDIATE ACTIONS")
    r += 1
    for i, action_txt in enumerate([nc.action_1, nc.action_2, nc.action_3], 1):
        ws.row_dimensions[r].height = 30
        cell(r, 2, i,
             fill=_fill(C_DARK_BLUE),
             font=_font(12, bold=True, color=C_WHITE),
             align=ALIGN_C,
             border=_border_all(C_DARK_BLUE))
        merge(r, 3, r, COL_E)
        cell(r, 3, action_txt or "",
             fill=_fill(C_WHITE),
             font=_font(10, color=C_TEXT_DARK),
             align=ALIGN_LT,
             border=_border_all(C_GRAY_LINE))
        r += 1

    # ── S7  ACCIONES CORRECTIVAS ──────────────────────────────────────────
    section_header(r, "  7.  ACCIONES CORRECTIVAS  /  CORRECTIVE ACTIONS")
    r += 1
    text_block(r, r + 3, nc.corrective_action)
    r += 4

    # ── S8  ESTADO DEL REPORTE ────────────────────────────────────────────
    section_header(r, "  8.  ESTADO DEL REPORTE")
    r += 1
    ws.row_dimensions[r].height = 15
    label_val(r, 2, 3, "ESTADO",       nc.status)
    label_val(r, 4, 5, "FECHA CIERRE", nc.closure_date.strftime(
        '%d/%m/%Y') if nc.closure_date else "")
    label_val(r, 6, 7, "CERRADA POR",
              nc.closed_by.username if nc.closed_by else "")
    for cc in range(8, 10):
        ws.cell(row=r, column=cc).fill = _fill(C_GRAY_BG)
        ws.cell(row=r, column=cc).border = _border_all(C_GRAY_LINE)
    r += 1

    # ── FOOTER ────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 4
    for c in range(COL_S, COL_E + 1):
        ws.cell(row=r, column=c).border = _border_top(C_ACCENT, "medium")
    r += 1

    ws.row_dimensions[r].height = 20
    merge(r, COL_S, r, COL_E)
    cell(r, COL_S,
         f"RS-SM-GC-REG-0001  |  Portal de Calidad  |  Grupo Riba Smith  |  "
         f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
         fill=_fill(C_DARK_BLUE),
         font=_font(10, bold=True, color=C_WHITE),
         align=ALIGN_C,
         border=_border_top(C_ACCENT, "medium"))

    return wb


# ══════════════════════════════════════════════════════════════════════════════
#  _build_list_workbook
# ══════════════════════════════════════════════════════════════════════════════

def _build_list_workbook(ncs):
    wb = Workbook()
    ws = wb.active
    ws.title = 'No Conformidades'
    ws.sheet_view.showGridLines = False

    brd = _border_all(C_GRAY_LINE)

    # Colores de estado y gravedad (unificados con paleta)
    GRAVITY_COLORS = {
        'Muy Alta': ("FEE2E2", "991B1B"),
        'Alta':     ("FEE2E2", "B91C1C"),
        'Media':    ("FEF3C7", "92400E"),
        'Baja':     ("DCFCE7", "166534"),
    }
    STATUS_COLORS = {
        'Abierta':  ("DBEAFE", "1D4ED8"),
        'Cerrada':  ("E2E8F0", "475569"),
        'Cerrado':  ("E2E8F0", "475569"),
    }

    COLS = [
        ('N° NC',            18), ('Fecha',          12), ('Hora',            8),
        ('Detectada',        14), ('Empresa',         20), ('T. Proveedor',   13),
        ('Sección',          20), ('Título',          40), ('Marca',          14),
        ('Causa',            14), ('Motivo',          12), ('Gravedad',       11),
        ('Evaluación',       18), ('Clasificación NC', 28), ('Cód. Ítem',      14),
        ('Producto NC',      24), ('Inv. Sucursal',   13), ('Valor Sucursal', 14),
        ('Unidades NC',      12), ('Total NC $',      12), ('Acción 1',       34),
        ('Acción 2',         34), ('Acción 3',
                                   34), ('Acción Correctiva', 40),
        ('Estado',           11), ('Fecha Cierre',    13),
    ]
    COLS = [
        ('NÂ° NC',            18), ('Fecha',          12), ('Hora',            8),
        ('Detectada',        14), ('Empresa',         20), ('T. Proveedor',   13),
        ('SecciÃ³n',          20), ('TÃ­tulo',          40), ('Marca',          14),
        ('Causa',            14), ('Motivo',          12), ('Gravedad',       11),
        ('EvaluaciÃ³n',       18), ('ClasificaciÃ³n NC', 28), ('CÃ³d. Ãtem',      14),
        ('CÃ³d. Barra',       18), ('Producto NC',     24), ('Inv. Sucursal',  13),
        ('Valor Sucursal',   14), ('Unidades NC',     12), ('Total NC $',     12),
        ('AcciÃ³n 1',         34), ('AcciÃ³n 2',        34), ('AcciÃ³n 3',       34),
        ('AcciÃ³n Correctiva', 40), ('Estado',         11), ('Fecha Cierre',   13),
    ]
    COLS = [
        ('Nro NC',             18), ('Fecha',            12), ('Hora',             8),
        ('Detectada',          14), ('Empresa',          20), ('T. Proveedor',    13),
        ('Seccion',            20), ('Titulo',           40), ('Marca',           14),
        ('Causa',              14), ('Motivo',           12), ('Gravedad',        11),
        ('Evaluacion',         18), ('Clasificacion NC', 28), ('Cod. Item',       14),
        ('Cod. Barra',         18), ('Producto NC',      24), ('Inv. Sucursal',   13),
        ('Valor Sucursal',     14), ('Unidades NC',      12), ('Total NC $',      12),
        ('Accion 1',           34), ('Accion 2',         34), ('Accion 3',        34),
        ('Accion Correctiva',  40), ('Estado',           11), ('Fecha Cierre',    13),
    ]
    n = len(COLS)
    last_col = get_column_letter(n)

    # ── ROW 1  Título ──────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].value = 'REPORTE DE NO CONFORMIDADES  /  GRUPO RIBA SMITH'
    ws['A1'].font = _font(14, bold=True, color=C_WHITE)
    ws['A1'].fill = _fill(C_DARK_BLUE)
    ws['A1'].alignment = ALIGN_C
    ws['A1'].border = _border_bottom(C_ACCENT, "medium")
    ws.row_dimensions[1].height = 30

    # ── ROW 2  Sub-header con conteo ──────────────────────────────────────
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'].value = (f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
                      f'  |  {len(ncs)} registros  |  Portal de Calidad')
    ws['A2'].font = _font(9, color="475569", italic=True)
    ws['A2'].fill = _fill(C_SOFT)
    ws['A2'].alignment = ALIGN_C
    ws['A2'].border = _border_bottom(C_GRAY_LINE)
    ws.row_dimensions[2].height = 14

    # ── ROW 3  Headers ────────────────────────────────────────────────────
    for ci, (lbl, w) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=ci, value=lbl)
        c.font = _font(10, bold=True, color=C_WHITE)
        c.fill = _fill(C_DARK_BLUE)
        c.alignment = ALIGN_C
        c.border = _border_bottom(C_ACCENT, "medium")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 24

    # Cols que van centradas
    CENTER_COLS = {1, 2, 3, 12, 26, 27}

    # ── ROWS de datos ─────────────────────────────────────────────────────
    for ri, nc in enumerate(ncs, 4):
        row_bg = "EEF2FF" if ri % 2 == 0 else C_WHITE   # alternancia suave azul/blanco

        row_vals = [
            nc.nc_number,
            nc.date_detected.strftime('%d/%m/%Y') if nc.date_detected else '',
            nc.time_detected.strftime('%H:%M') if nc.time_detected else '',
            nc.detected_by_type or '',
            nc.empresa or '',
            nc.provider_tipo or '',
            nc.section or '',
            nc.title,
            nc.marca or '',
            nc.cause or '',
            nc.motive or '',
            nc.gravity or '',
            nc.evaluations or '',
            nc.classification or '',
            nc.product_code or '',
            nc.barcode or '',
            nc.product_affected or '',
            nc.branch_total_units or '',
            nc.branch_total_usd or '',
            nc.total_units or '',
            nc.total_usd or '',
            nc.action_1 or '',
            nc.action_2 or '',
            nc.action_3 or '',
            nc.corrective_action or '',
            nc.status,
            nc.closure_date.strftime('%d/%m/%Y') if nc.closure_date else '',
        ]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = brd
            c.alignment = ALIGN_C if ci in CENTER_COLS else ALIGN_L

            if ci == 12 and nc.gravity in GRAVITY_COLORS:
                bg_col, txt_col = GRAVITY_COLORS[nc.gravity]
                c.fill = _fill(bg_col)
                c.font = _font(9, bold=True, color=txt_col)
            elif ci == 26 and nc.status in STATUS_COLORS:
                bg_col, txt_col = STATUS_COLORS[nc.status]
                c.fill = _fill(bg_col)
                c.font = _font(9, bold=True, color=txt_col)
            elif ci == 1:
                # N° NC destacado
                c.fill = _fill(C_SOFT)
                c.font = _font(9, bold=True, color=C_DARK_BLUE)
            else:
                c.fill = _fill(row_bg)
                c.font = _font(9, color=C_TEXT_DARK)

        ws.row_dimensions[ri].height = 28

    # ── Freeze + filtros ──────────────────────────────────────────────────
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{last_col}3'

    # ── FOOTER ────────────────────────────────────────────────────────────
    footer_row = len(ncs) + 4
    ws.row_dimensions[footer_row].height = 4
    for c in range(1, n + 1):
        ws.cell(row=footer_row, column=c).border = _border_top(
            C_ACCENT, "medium")

    footer_row += 1
    ws.merge_cells(f'A{footer_row}:{last_col}{footer_row}')
    fc = ws.cell(row=footer_row, column=1,
                 value=f"RS-SM-GC-REG-0001  |  Portal de Calidad  |  Grupo Riba Smith  |  "
                 f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    fc.fill = _fill(C_DARK_BLUE)
    fc.font = _font(10, bold=True, color=C_WHITE)
    fc.alignment = ALIGN_C
    fc.border = _border_top(C_ACCENT, "medium")
    ws.row_dimensions[footer_row].height = 20

    return wb
